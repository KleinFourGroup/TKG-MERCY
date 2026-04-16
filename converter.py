import openpyxl
import re

from records import Material, ImportedPart, Part, Package, Mixture, Globals, Database

def resplit(delimiters, string, maxsplit=0):
    regex_pattern = '|'.join(map(re.escape, delimiters))
    return re.split(regex_pattern, string, maxsplit)

def numlist(string):
    regex_pattern = '[^\\d][^\\d]*'
    tokens = re.split(regex_pattern, string)
    return [int(num) for num in tokens if not num == ""]

def getString(ws, col, row):
    cell = ws[col + str(row)]
    val = cell.value
    if val == None:
        val = ""
    assert(isinstance(val, str))
    return val

def getStringForce(ws, col, row):
    cell = ws[col + str(row)]
    val = cell.value
    if val == None:
        val = ""
    return str(val)

def getNumber(ws, col, row):
    cell = ws[col + str(row)]
    val = cell.value
    if val == None:
        val = 0
    if not (isinstance(val, int) or isinstance(val, float)):
        print("Error at {}{}: expected number, got {} ({})".format(col, row, val, type(val)))
    assert(isinstance(val, int) or isinstance(val, float))
    return val

def getMaterial(mats, row):
    mat = Material(getString(mats, "B", row))
    mat.setChems(getNumber(mats, "F", row), getNumber(mats, "G", row), getNumber(mats, "H", row), getNumber(mats, "I", row), getNumber(mats, "J", row), getNumber(mats, "K", row), getNumber(mats, "L", row), getNumber(mats, "M", row), getNumber(mats, "N", row), getNumber(mats, "O", row), getNumber(mats, "P", row))
    mat.setSizes(getNumber(mats, "R", row), getNumber(mats, "S", row), getNumber(mats, "T", row), getNumber(mats, "U", row), getNumber(mats, "V", row))
    return mat

def importMaterials():
    matb = openpyxl.load_workbook("Chemistry and Sizing Worksheet.xlsx", data_only=True)
    mats = matb.active
    res = {}
    for row in range(5, 84, 2):
        mat = getMaterial(mats, row)
        res[mat.name] = mat
    matb.close()
    return res



def getPackaging(packs, part, row):
    pads = resplit([", ", "/"], getString(packs, "J", row))
    rawcounts = packs["K" + str(row)].value
    if isinstance(rawcounts, float):
        rawcounts = int(rawcounts)
    rawcounts = str(rawcounts)
    padcounts = list(map(int, rawcounts.split(",")))
    palletRow = int(getStringForce(packs, "I", row)[4:])
    miscRows = numlist(getStringForce(packs, "N", row))
    miscs = [getString(packs, "J", mrow) for mrow in miscRows]
    part.setPackaging(getString(packs, "B", row), int(getNumber(packs, "C", row)), getString(packs, "G", palletRow), int(getNumber(packs, "D", row)), pads, padcounts, miscs)

def getPart(parts, parts_form, packs, row):
    part = ImportedPart(getString(parts, "A", row))
    try:
        part.setProduction(getNumber(parts, "B", row), int(getNumber(parts, "C", row)), getStringForce(parts, "D", row), getNumber(parts, "J", row), getNumber(parts, "L", row), getNumber(parts, "N", row), getNumber(parts, "P", row), getNumber(parts, "R", row), getNumber(parts, "V", row), getNumber(parts, "AF", row))
    except:
        return None
    try:
        getPackaging(packs, part, int(getString(parts_form, "X", row)[12:]))
    except:
        return None
    return part


def importParts():
    prodb = openpyxl.load_workbook("Product Costing 2024A.xlsx", data_only=True)
    prodb_form = openpyxl.load_workbook("Product Costing 2024A.xlsx")
    parts = prodb["Part Costs"]
    parts_form = prodb_form["Part Costs"]
    packs = prodb_form["Packaging"]
    impRes = {}
    res = {}
    rows = {}
    for row in range(5, 94):
        if not (parts["C" + str(row)].value == None or parts["L" + str(row)].value == None or parts["L" + str(row)].value == 0 or parts["R" + str(row)].value == None):
            part = getPart(parts, parts_form, packs, row)
            if not part == None:
                impRes[part.name] = part
                res[part.name] = part.convert()
                rows[part.name] = row
            else:
                print("Error with row {}".format(row))
    prodb.close()
    prodb_form.close()
    return res, impRes, rows



def importPackaging():
    prodb = openpyxl.load_workbook("Product Costing 2024A.xlsx", data_only=True)
    packs = prodb["Packaging"]
    res = {}
    for row in range(4, 44, 2):
        pack = Package(getString(packs, "A", row), "box", getNumber(packs, "B", row))
        res[pack.name] = pack
    for row in range(4, 36, 2):
        pack = Package(getString(packs, "D", row), "pad", getNumber(packs, "E", row))
        res[pack.name] = pack
    for row in range(4, 6, 2):
        pack = Package(getString(packs, "G", row), "pallet", getNumber(packs, "H", row))
        res[pack.name] = pack
    for row in range(4, 22, 2):
        pack = Package(getString(packs, "J", row), "misc", getNumber(packs, "K", row))
        res[pack.name] = pack
    prodb.close()
    return res



def importMixes(materials):
    prodb = openpyxl.load_workbook("Product Costing 2024A.xlsx", data_only=True)
    mixs = prodb["Mix Cost"]
    res = {}
    row = 2
    while row > 0:
        mix = Mixture(getStringForce(mixs, "B", row))
        row +=3
        while True:
            mat = getString(mixs, "A", row)
            if mat == "Total":
                break
            else:
                mix.add(mat, getNumber(mixs, "B", row))
                if mat in materials:
                    materials[mat].setCost(getNumber(mixs, "E", row), getNumber(mixs, "F", row))
                else:
                    print("Unknown material {}".format(mat))
                row += 2
        res[mix.name] = mix
        row += 1
        while True:
            if not getString(mixs, "A", row) == "":
                row = 0
                break
            if not getStringForce(mixs, "B", row) == "":
                break
            row += 1
    prodb.close()
    return res

def check(name, db: Database, checkRows, parts):
    part = db.parts[name]
    row = checkRows[name]
    err = 0
    err += abs(part.getMatlCost() - getNumber(parts, "G", row))
    err += abs(part.getBatchingTime() - getNumber(parts, "I", row))
    err += abs(part.getPressingTime() - getNumber(parts, "K", row))
    err += abs(part.getTurningTime() - getNumber(parts, "M", row))
    # err += abs(part.getLoadingTime() - getNumber(parts, "O", row))
    # err += abs(part.getUnloadingTime() - getNumber(parts, "Q", row))
    # err += abs(part.getInspectionTime() - getNumber(parts, "S", row))
    err += abs(part.getLaborCost() - getNumber(parts, "U", row))
    err += abs(part.getGrossMatlLaborCost() - getNumber(parts, "W", row))
    err += abs(part.getPackagingCost() - getNumber(parts, "X", row))
    err += abs(part.getManufacturingOverhead() - getNumber(parts, "Y", row))
    err += abs(part.getManufacturingCost() - getNumber(parts, "AA", row))
    err += abs(part.getSGA() - getNumber(parts, "AC", row))
    err += abs(part.getTotalCost() - getNumber(parts, "AE", row))
    err += abs(part.getGM() - getNumber(parts, "AG", row))
    err += abs(part.getCM() - getNumber(parts, "AH", row))
    err += abs(part.getVariableCost() - getNumber(parts, "AI", row))
    err += abs(part.getProductivity() - getNumber(parts, "AK", row))
    return err

def importDatabase(checkF = False):
    materials = importMaterials()
    parts, importedParts, checkRows = importParts()
    packaging = importPackaging()
    mixes = importMixes(materials)
    db = Database(Globals(), materials, mixes, packaging, parts)
    if checkF:
        impDb = Database(Globals(), materials, mixes, packaging, importedParts)
        prodb = openpyxl.load_workbook("Product Costing 2024A.xlsx", data_only=True)
        parts = prodb["Part Costs"]
        for entry in db.parts:
            err = check(entry, db, checkRows, parts)
            errImp = check(entry, impDb, checkRows, parts)
            print("{} | {:.4f} {:.4f}".format(entry, abs(err - errImp), err))
        prodb.close()
    return db
